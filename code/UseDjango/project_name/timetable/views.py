from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.cache import cache
from django.db.models import Count
from datetime import datetime, timedelta
from collections import defaultdict
import json
import logging

logger = logging.getLogger(__name__)

# Export libraries
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from xml.sax.saxutils import escape

# Model imports
from .models import (
    setupModel, yearSetupModel, subjectModel, practicalModel,
    yearClassRoomModel, yearLabRoomModel, DepartmentTimetableModel
)

# Form imports
from .forms import (
    setup_form, TeacherFormSet, ClassRoomFormSet, LabFormSet, RecessFormSet,
    yearSetupForm, SubjectForm, PracticalForm, YearClassRoomForm, YearLabRoomForm,
    EditYearStudentsInBatchForm, EditYearTotalStudentsForm
)

# Service imports
from timetable.services.schedule import run_full_timetable_scheduler

from timetable.services.data_utils import (
    DAY_NAMES,
    calculate_1or_2hour_slots, createData,
    generate_time_range, is_recess_time, normalize_recess_windows,
    combine_timetables, prepare_rows,
    calculate_hours_per_week, prepare_lecture_hours_table,
    prepare_practical_hours_table, prepare_teacher_vs_year_chart,
)



def home(request):
    stats = {
        'departments': setupModel.objects.count(),
        'years': yearSetupModel.objects.count(),
        'subjects': subjectModel.objects.count(),
        'practicals': practicalModel.objects.count(),
        'saved_timetables': DepartmentTimetableModel.objects.count(),
    }

    recent_departments = setupModel.objects.annotate(
        years_total=Count('relatedNameYearSetupModel', distinct=True),
        recess_total=Count('relatedNameRecesses', distinct=True),
        saved_tt_total=Count('realatedNameSavedTimetablesOfDepartment', distinct=True),
    ).order_by('-id')[:4]

    recent_timetables = DepartmentTimetableModel.objects.select_related(
        'department'
    ).order_by('-created_at')[:5]

    return render(request, 'timetable/home.html', {
        'stats': stats,
        'recent_departments': recent_departments,
        'recent_timetables': recent_timetables,
    })


def setup(request):
    setup_objects = setupModel.objects.prefetch_related('relatedNameYearSetupModel',
     'relatedNameRecesses',
    )

    for department in setup_objects:
        # building recess list
        recess_list = [
            [
                r.recess_start_time.strftime("%H:%M"),
                r.recess_end_time.strftime("%H:%M")
            ]
            for r in department.relatedNameRecesses.all()
        ]

        depStartTime = department.start_time.strftime("%H:%M")
        depEndTime = department.end_time.strftime("%H:%M")

        # calculating lecture and practical time slots to display in years page once
        department.all_lecture_timeslots = calculate_1or_2hour_slots(
            department.hours_in_lecture, depStartTime, depEndTime, recess_list
        )

        department.all_practical_timeslots = calculate_1or_2hour_slots(
            department.hours_in_practical, depStartTime, depEndTime, recess_list
        )
    return render(request, 'timetable/setup.html',{

        'setup_objects': setup_objects,
    })


# viewing the setupForm
def setupForm(request, pk=None):

    setup_instance = get_object_or_404(setupModel, pk=pk) if pk else None

#     # if the method is post then we are getting the data from the form i.e submitting the form

    if request.method == "POST":
        form1 = setup_form(request.POST, instance=setup_instance)
        # now the inline formsets
        if setup_instance:
            formset1 = TeacherFormSet(request.POST, instance=setup_instance)
            formset2 = ClassRoomFormSet(request.POST, instance=setup_instance)
            formset3 = LabFormSet(request.POST, instance=setup_instance)
            formset4 = RecessFormSet(request.POST, instance=setup_instance)
        else:
            formset1 = TeacherFormSet(request.POST)
            formset2 = ClassRoomFormSet(request.POST)
            formset3 = LabFormSet(request.POST)
            formset4 = RecessFormSet(request.POST)

        if form1.is_valid() and formset1.is_valid() and formset2.is_valid() and formset3.is_valid() and formset4.is_valid():

            # saving the form data to database and getting the instance of setupModel
            setup_instance = form1.save()
            # now saving the formsets
            formset1.instance = setup_instance
            formset1.save()
            formset2.instance = setup_instance
            formset2.save()
            formset3.instance = setup_instance
            formset3.save()
            formset4.instance = setup_instance
            formset4.save()

            if pk:
                messages.success(request, 'Department updated successfully.')
            else:
                messages.success(request, 'Setup data saved successfully.')

            # redirecting to setup page after saving the data
            return redirect('timetable:setup')
    else:
        form1 = setup_form(instance=setup_instance)
        if setup_instance:
            formset1 = TeacherFormSet(instance=setup_instance)
            formset2 = ClassRoomFormSet(instance=setup_instance)
            formset3 = LabFormSet(instance=setup_instance)
            formset4 = RecessFormSet(instance=setup_instance)
        else:
            formset1 = TeacherFormSet()
            formset2 = ClassRoomFormSet()
            formset3 = LabFormSet()
            formset4 = RecessFormSet()

    return render(request, 'timetable/setupform.html', {
        'form': form1,
        'formset1': formset1,
        'formset2': formset2,
        'formset3': formset3,
        'formset4': formset4,
        'editing': bool(pk),
        'setup_instance': setup_instance,

        })



# setup delete button for deleting the entire departmet stuff
def setup_delete(request, pk):
    if request.method == "POST":
        setup_object = get_object_or_404(setupModel, pk=pk)
        setup_object.delete()
        messages.success(request, "Department deleted successfully.")
    return redirect('timetable:setup')  # ← stays on same page

# for viewing years of that department, and main display
def years(request, pk):
    department = get_object_or_404(setupModel, pk=pk)
    
    # Track last visited department in session for better navigation
    request.session['last_department_id'] = department.id
    request.session['last_department_name'] = department.department_name
    
    # Optimized query with prefetch_related to avoid N+1 queries
    years_of_department = department.relatedNameYearSetupModel.prefetch_related(
        'subjects',
        'subjects__teachers',
        'practicals',
        'practicals__teachers',
        'yearClassRoomsAllocated',
        'yearLabRoomsAllocated'
    )
    
    # checking if department has years or not
    has_years = years_of_department.exists()
    # calculating lecture and practical time slots to display in years page
    prac_hours = department.hours_in_practical
    lec_hours = department.hours_in_lecture
    department_starttime = department.start_time.strftime("%H:%M")
    department_endtime = department.end_time.strftime("%H:%M")
    recess_list = []
    total_recessHoursInDay = 0
    for each_recess in department.relatedNameRecesses.all():
                recess_list.append([
                    each_recess.recess_start_time.strftime("%H:%M"),
                    each_recess.recess_end_time.strftime("%H:%M")
                ])
                # calculating total recess hours in a day
                # we can substract date time therefore we have to convert to date of today with recess start time/end time
                start_date = datetime.combine(datetime.today(), each_recess.recess_start_time)
                end_date = datetime.combine(datetime.today(), each_recess.recess_end_time)

                eachrecesshour = (end_date - start_date).total_seconds() / 3600
                total_recessHoursInDay += eachrecesshour

    practical_timeslots = calculate_1or_2hour_slots(prac_hours, department_starttime, department_endtime, recess_list)
    lecture_timeslots = calculate_1or_2hour_slots(lec_hours, department_starttime, department_endtime, recess_list)


    # THESE are same for each year of a department as we take start time and end time from department

    # total working hours in a day  calculation
    total_workingHoursInDay = (datetime.strptime(department_endtime, "%H:%M") - datetime.strptime(department_starttime, "%H:%M")).total_seconds()/3600
    # logger.debug(total_workingHoursInDay)

    # logger.debug(total_recessHoursInDay)


    total_workingHoursInWeek = total_workingHoursInDay * department.number_of_days
    total_recessHoursInWeek = total_recessHoursInDay * department.number_of_days

    # for each year
    for eachYear in years_of_department:
        # chart prep
        eachYear.donutChart1_data = {}

        # calculating total lecture hours per week for that year
        sumLectureHoursPerWeek = 0
        for oneSubject in eachYear.subjects.all():
            sumLectureHoursPerWeek += oneSubject.hours_per_week 
            

        eachYear.donutChart1_data['totalLectureHoursPerWeek'] = sumLectureHoursPerWeek
        # logger.debug(sumLectureHoursPerWeek)

        # calculating total practical hours per week for that year
        sumPracticalHoursPerWeek = 0
        for onePractical in eachYear.practicals.all():
            sumPracticalHoursPerWeek += onePractical.hours_per_week
        
        eachYear.donutChart1_data['totalPracticalHoursPerWeek'] = sumPracticalHoursPerWeek
        # logger.debug("practical sum",sumPracticalHoursPerWeek)

        # calculating free hours per week for that year

        unAllocatedHoursPerWeek = total_workingHoursInWeek-(sumLectureHoursPerWeek + sumPracticalHoursPerWeek + total_recessHoursInWeek)
        eachYear.donutChart1_data['unAllocatedHoursPerWeek'] = unAllocatedHoursPerWeek



    # donut chart to display



    


    



#  calculate_1or_2hour_slots(pracHOURS, departmentSTARTTIME, departmentENDTIME, recess_list  )

    return render(request, 'timetable/years.html',{
        'department': department,
        'years_of_department': years_of_department,
        'practical_timeslots': practical_timeslots,
        'lecture_timeslots': lecture_timeslots,
        'has_years': has_years,

        'total_recessHoursInWeek': total_recessHoursInWeek,
        'total_workingHoursInWeek': total_workingHoursInWeek,
    }
    )
# for taking main years data, form
def yearsSetupForm(request, department_id):
    # take the department 
    department = get_object_or_404(setupModel, pk=department_id)
    if request.method == "POST":
        form = yearSetupForm(request.POST)

        if form.is_valid():

            year = form.save(commit=False)
            year.department = department
            year.save()
            messages.success(request, "Year created successfully.")

            return redirect('timetable:years', pk=department.id)
    else:
        form = yearSetupForm()

    return render(request, 'timetable/yearsform.html', {
        'form': form,
        'department': department,
    })


def add_subject(request, department_id, year_id):
    department = get_object_or_404(setupModel, pk=department_id)
    year = get_object_or_404(yearSetupModel, pk=year_id, department=department)

    if request.method == "POST":
        form = SubjectForm(request.POST, department=department, year=year)
        if form.is_valid():
            subject = form.save(commit=False)
            subject.year = year
            subject.save()
            form.save_m2m()  # VERY IMPORTANT for ManyToMany
            messages.success(request, "Subject added successfully.")

            return redirect('timetable:years', pk=department.id)
    else:
        form = SubjectForm(department=department, year=year)

    return render(request, 'timetable/add_subject.html', {
        'form': form,
        'year': year,
        'department': department
    })

def delete_year(request, department_id, year_id):
    department = get_object_or_404(setupModel, pk=department_id)
    year = get_object_or_404(
        yearSetupModel,
        pk=year_id,
        department=department
    )

    if request.method == "POST":
        year.delete()
        messages.success(request, "Year deleted successfully.")

    return redirect('timetable:years', pk=department.id)


def delete_subject(request, department_id, subject_id):
    department = get_object_or_404(setupModel, pk=department_id)

    subject = get_object_or_404(
        subjectModel,
        pk=subject_id,
        year__department=department
    )

    if request.method == "POST":
        subject.delete()
        messages.success(request, "Subject deleted successfully.")

    return redirect('timetable:years', pk=department.id)

def edit_subject(request, department_id, subject_id):
    department = get_object_or_404(setupModel, pk=department_id)
    subject = get_object_or_404(
        subjectModel,
        pk=subject_id,
        year__department=department
    )
    year = subject.year

    if request.method == "POST":
        form = SubjectForm(request.POST, instance=subject, department=department, year=year)
        if form.is_valid():
            form.save()
            messages.success(request, "Subject updated successfully.")
            return redirect('timetable:years', pk=department.id)
    else:
        form = SubjectForm(instance=subject, department=department, year=year)

    return render(request, 'timetable/add_subject.html', {
        'form': form,
        'year': year,
        'department': department,
        'editing': True,
        'subject': subject,
    })

def add_practical(request, department_id, year_id):
    department = get_object_or_404(setupModel, pk=department_id)
    year = get_object_or_404(yearSetupModel, pk=year_id, department=department)

    if request.method == "POST":
        form = PracticalForm(request.POST, department=department, year=year)
        if form.is_valid():
            practical = form.save(commit=False)
            practical.year = year
            practical.save()
            form.save_m2m()
            messages.success(request, "Practical added successfully.")

            return redirect('timetable:years', pk=department.id)
    else:
        form = PracticalForm(department=department, year=year)

    return render(request, 'timetable/add_practical.html', {
        'form': form,
        'year': year,
        'department': department
    })

def delete_practical(request, department_id, practical_id):
    department = get_object_or_404(setupModel, pk=department_id)
    practical = get_object_or_404(
        practicalModel,
        pk=practical_id,
        year__department=department
    )

    if request.method == "POST":
        practical.delete()
        messages.success(request, "Practical deleted successfully.")

    return redirect('timetable:years', pk=department.id)

def edit_practical(request, department_id, practical_id):
    department = get_object_or_404(setupModel, pk=department_id)
    practical = get_object_or_404(
        practicalModel,
        pk=practical_id,
        year__department=department
    )
    year = practical.year

    if request.method == "POST":
        form = PracticalForm(request.POST, instance=practical, department=department, year=year)
        if form.is_valid():
            form.save()
            messages.success(request, "Practical updated successfully.")
            return redirect('timetable:years', pk=department.id)
    else:
        form = PracticalForm(instance=practical, department=department, year=year)

    return render(request, 'timetable/add_practical.html', {
        'form': form,
        'year': year,
        'department': department,
        'editing': True,
        'practical': practical,
    })

def add_year_classrooms(request, department_id, year_id):
    department = get_object_or_404(setupModel, pk=department_id)
    year = get_object_or_404(yearSetupModel, pk=year_id, department=department)

    allocation, _ = yearClassRoomModel.objects.get_or_create(year=year)

    if request.method == "POST":
        form = YearClassRoomForm(
            request.POST,
            instance=allocation,
            department=department
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Classrooms allocated successfully.")
            return redirect('timetable:years', pk=department.id)
    else:
        form = YearClassRoomForm(
            instance=allocation,
            department=department
        )

    return render(request, 'timetable/add_year_classrooms.html', {
        'form': form,
        'year': year,
        'department': department
    })

def add_year_labrooms(request, department_id, year_id):
    department = get_object_or_404(setupModel, pk=department_id)
    year = get_object_or_404(yearSetupModel, pk=year_id, department=department)

    allocation, _ = yearLabRoomModel.objects.get_or_create(year=year)

    if request.method == "POST":
        form = YearLabRoomForm(
            request.POST,
            instance=allocation,
            department=department
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Lab rooms allocated successfully.")
            return redirect('timetable:years', pk=department.id)
    else:
        form = YearLabRoomForm(
            instance=allocation,
            department=department
        )

    return render(request, 'timetable/add_year_labrooms.html', {
        'form': form,
        'year': year,
        'department': department
    })

# to edit number of students in each batch for a year
def edit_year_students_in_batch(request, department_id, year_id):
    department = get_object_or_404(setupModel, pk=department_id)
    year = get_object_or_404(yearSetupModel, pk=year_id, department=department)

    if request.method == "POST":
        form = EditYearStudentsInBatchForm(request.POST, instance=year)
        if form.is_valid():
            form.save()
            messages.success(request, "Students per batch updated successfully.")
            return redirect('timetable:years', pk=department.id)
    else:
        form = EditYearStudentsInBatchForm(instance=year)
    
    return render(request, 'timetable/edit_year_students_in_batch.html', {
        'form': form,
        'year': year,
        'department': department
    })


# to edit total number of students for a year
def edit_year_total_students(request, department_id, year_id):
    department = get_object_or_404(setupModel, pk=department_id)
    year = get_object_or_404(yearSetupModel, pk=year_id, department=department)

    if request.method == "POST":
        form = EditYearTotalStudentsForm(request.POST, instance=year)
        if form.is_valid():
            form.save()
            messages.success(request, "Total students updated successfully.")
            return redirect('timetable:years', pk=department.id)
    else:
        form = EditYearTotalStudentsForm(instance=year)
    
    return render(request, 'timetable/edit_year_total_students.html', {
        'form': form,
        'year': year,
        'department': department
    })


def validate_department_for_generation(department):
    """
    Check that a department is fully configured before running the scheduler.
    Returns a list of human-readable error strings. Empty list = ready to generate.
    """
    errors = []

    years = department.relatedNameYearSetupModel.prefetch_related(
        'subjects', 'subjects__teachers',
        'practicals', 'practicals__teachers',
        'yearClassRoomsAllocated', 'yearLabRoomsAllocated',
    ).all()

    if not years.exists():
        errors.append("The department has no years configured. Add at least one year.")
        return errors  # no point checking further

    for year in years:
        name = year.year_name

        # ── Subjects ────────────────────────────────────────
        subjects = list(year.subjects.all())
        if not subjects:
            errors.append(f"[{name}] No subjects added. Add at least one subject.")
        else:
            for subj in subjects:
                if not subj.teachers.exists():
                    errors.append(f"[{name}] Subject '{subj.subject_name}' has no teacher assigned.")

        # ── Classrooms ──────────────────────────────────────
        try:
            classroom_alloc = year.yearClassRoomsAllocated
            if not classroom_alloc.classrooms.exists():
                errors.append(f"[{name}] No classrooms allocated. Allocate at least one classroom.")
        except Exception:
            errors.append(f"[{name}] No classrooms allocated. Allocate at least one classroom.")

        # ── Practicals (optional — only validate if any exist) ──
        practicals = list(year.practicals.all())
        if practicals:
            for prac in practicals:
                if not prac.teachers.exists():
                    errors.append(f"[{name}] Practical '{prac.practical_name}' has no teacher assigned.")

            try:
                labroom_alloc = year.yearLabRoomsAllocated
                if not labroom_alloc.labrooms.exists():
                    errors.append(f"[{name}] No lab rooms allocated, but practicals are defined.")
            except Exception:
                errors.append(f"[{name}] No lab rooms allocated, but practicals are defined.")

        # ── Student count ────────────────────────────────────
        if year.total_students <= 0:
            errors.append(f"[{name}] Total students must be greater than 0.")
        if practicals and year.number_of_students_in_batch <= 0:
            errors.append(f"[{name}] Students per batch must be greater than 0.")

    return errors


def generate_timetableData(request, department_id):

    department = get_object_or_404(setupModel, pk=department_id)

    # =========================
    # HANDLE SAVE FIRST
    # =========================
    if request.method == "POST" and ("save" in request.POST or "name" in request.POST):

        name = (request.POST.get("name") or "").strip()
        if not name:
            messages.error(request, "Please enter a timetable name before saving.")
            return redirect("timetable:generate_timetable", department_id=department.id)

        #  We must NOT regenerate.
        # So we fetch already generated data from session.
        display_tt = request.session.get("generated_display_tt")
        fitness = request.session.get("generated_fitness")
        teacher_year_chart = request.session.get("generated_teacher_chart")
        lecture_hours_table = request.session.get("generated_lecture_table")
        practical_hours_table = request.session.get("generated_practical_table")

        # Session can be missing if page was refreshed or session expired.
        # In that case regenerate payload once and continue saving.
        if not display_tt:
            validation_errors = validate_department_for_generation(department)
            if validation_errors:
                return render(request, "timetable/error.html", {
                    "error": " | ".join(validation_errors),
                    "department": department,
                    "validation_errors": validation_errors,
                })

            year_data_all = createData(department)
            timeslot_ranges = generate_time_range(
                year_data_all[next(iter(year_data_all))]['each_lectureHours'],
                year_data_all[next(iter(year_data_all))]['start_time_str'],
                year_data_all[next(iter(year_data_all))]['end_time_str'],
                year_data_all[next(iter(year_data_all))]['recess_list']
            )
            result = run_full_timetable_scheduler(year_data_all)

            if not result["success"]:
                return render(request, "timetable/error.html", {
                    "error": result["error"],
                    "department": department,
                })

            combined_timetables = combine_timetables(
                result["lecture_timetable"],
                result["practical_timetable"],
                year_data_all
            )

            hours_stats = calculate_hours_per_week(combined_timetables, year_data_all)
            teacher_year_chart = prepare_teacher_vs_year_chart(hours_stats) or {"teachers": [], "years": {}}
            lecture_hours_table = prepare_lecture_hours_table(hours_stats, year_data_all)
            practical_hours_table = prepare_practical_hours_table(hours_stats, year_data_all)
            display_tt = prepare_rows(combined_timetables, timeslot_ranges)
            fitness = result.get("fitness", 0)

            request.session["generated_display_tt"] = display_tt
            request.session["generated_fitness"] = fitness
            request.session["generated_teacher_chart"] = teacher_year_chart
            request.session["generated_lecture_table"] = lecture_hours_table
            request.session["generated_practical_table"] = practical_hours_table

        try:
            DepartmentTimetableModel.objects.create(
                department=department,
                name=name,
                displayTimetable=display_tt,
                fitness=fitness,
                teacherYearChart=teacher_year_chart,
                lectureHoursTable=lecture_hours_table,
                practicalHoursTable=practical_hours_table,
            )
        except Exception:
            logger.exception("Failed to save timetable", extra={"department_id": department.id, "name": name})
            messages.error(request, "Could not save timetable. Please try again.")
            return redirect("timetable:generate_timetable", department_id=department.id)

        messages.success(request, "Timetable saved successfully.")

        return redirect(
            "timetable:department_saved_timetable_list",
            department_id=department.id
        )

    # =========================
    # ONLY GENERATE ON GET
    # =========================

    # Pre-flight validation — show friendly errors before touching the scheduler
    validation_errors = validate_department_for_generation(department)
    if validation_errors:
        return render(request, "timetable/error.html", {
            "error": " | ".join(validation_errors),
            "department": department,
            "validation_errors": validation_errors,
        })

    year_data_all = createData(department)

    timeslot_ranges = generate_time_range(
        year_data_all[next(iter(year_data_all))]['each_lectureHours'],
        year_data_all[next(iter(year_data_all))]['start_time_str'],
        year_data_all[next(iter(year_data_all))]['end_time_str'],
        year_data_all[next(iter(year_data_all))]['recess_list']
    )

    result = run_full_timetable_scheduler(year_data_all)

    if not result["success"]:
        return render(request, "timetable/error.html", {
            "error": result["error"],
            "department": department,
        })

    combined_timetables = combine_timetables(
        result["lecture_timetable"],
        result["practical_timetable"],
        year_data_all
    )

    hours_stats = calculate_hours_per_week(combined_timetables, year_data_all)

    teacher_year_chart = prepare_teacher_vs_year_chart(hours_stats)
    # Ensure teacher_year_chart has valid structure
    if not teacher_year_chart:
        teacher_year_chart = {"teachers": [], "years": {}}
    lecture_hours_table = prepare_lecture_hours_table(hours_stats, year_data_all)
    practical_hours_table = prepare_practical_hours_table(hours_stats, year_data_all)

    display_tt = prepare_rows(combined_timetables, timeslot_ranges)
    fitness = result.get("fitness", 0)

    # 🔥 STORE IN SESSION
    request.session["generated_display_tt"] = display_tt
    request.session["generated_fitness"] = fitness
    request.session["generated_teacher_chart"] = teacher_year_chart
    request.session["generated_lecture_table"] = lecture_hours_table
    request.session["generated_practical_table"] = practical_hours_table

    return render(request, "timetable/generate.html", {
        'department': department,
        "display_tt": display_tt,
        "fitness": fitness,
        "teacher_year_chart_data": teacher_year_chart,
        "teacher_year_chart": json.dumps(teacher_year_chart),
        "lecture_hours_table": lecture_hours_table,
        "practical_hours_table": practical_hours_table,
    })

# for viewing the list of saved timetables

def department_saved_timetable_list(request, department_id):
    department = get_object_or_404(setupModel, pk=department_id)

    # Optimized query with select_related for department
    saved_tts_queryset = department.realatedNameSavedTimetablesOfDepartment.select_related(
        'department'
    ).order_by('-created_at')
    
    # Pagination for performance with large datasets
    paginator = Paginator(saved_tts_queryset, 20)  # Show 20 timetables per page
    page = request.GET.get('page', 1)
    
    try:
        saved_tts = paginator.page(page)
    except PageNotAnInteger:
        saved_tts = paginator.page(1)
    except EmptyPage:
        saved_tts = paginator.page(paginator.num_pages)

    return render(request, "timetable/saved_timetable_list.html", {
        'department': department,
        'saved_tts': saved_tts,
        'is_paginated': paginator.num_pages > 1,
        'page_obj': saved_tts,
    })

# for viewing saved timetables
def view_saved_timetable(request, tt_id):
    tt = get_object_or_404(DepartmentTimetableModel, pk=tt_id)
    tt_name = tt.name

    display_tt = tt.displayTimetable
    department_recess_starts = sorted([
        r.recess_start_time.strftime("%H:%M")
        for r in tt.department.relatedNameRecesses.all()
    ])

    # fitness
    fitness = tt.fitness
    # chart data
    teacher_year_chart = tt.teacherYearChart
    # Ensure teacher_year_chart has valid structure
    if not teacher_year_chart:
        teacher_year_chart = {"teachers": [], "years": {}}
    lecture_hours_table = tt.lectureHoursTable
    practical_hours_table = tt.practicalHoursTable


    return render(request, "timetable/view_saved_timetable.html", {
        'department': tt.department,
        # "practical_timetable": result["practical_timetable"],
        # "lecture_timetable": result["lecture_timetable"],
        "tt_id": tt_id,
        "tt_name": tt_name,
        "fitness": fitness,
        # "combined_timetable": combined_timetables,
        # "timeslot_ranges": timeslot_ranges,
        "display_tt": display_tt,
        
        # for chart and table data
        # "hours_stats": hours_stats,
        "teacher_year_chart": json.dumps(teacher_year_chart),
        "lecture_hours_table": lecture_hours_table,
        "practical_hours_table": practical_hours_table,
    })


def delete_saved_timetable(request, tt_id):
    tt = get_object_or_404(DepartmentTimetableModel, pk=tt_id)
    department_id = tt.department.id
    tt.delete()

    return redirect("timetable:department_saved_timetable_list", department_id=department_id)


def compare_timetables(request, department_id):
    """Compare multiple timetables side-by-side"""
    department = get_object_or_404(setupModel, pk=department_id)
    
    # Get selected timetable IDs from query parameters
    tt_ids = request.GET.getlist('tt_ids')
    
    if not tt_ids:
        messages.error(request, 'Please select at least one timetable to compare.')
        return redirect('timetable:department_saved_timetable_list', department_id=department_id)
    
    if len(tt_ids) > 3:
        messages.warning(request, 'You can compare up to 3 timetables at once. Showing first 3.')
        tt_ids = tt_ids[:3]
    
    # Fetch timetables
    timetables = []
    for tt_id in tt_ids:
        try:
            tt = DepartmentTimetableModel.objects.get(pk=int(tt_id), department=department)
            timetables.append(tt)
        except (DepartmentTimetableModel.DoesNotExist, ValueError):
            continue
    
    if not timetables:
        messages.error(request, 'No valid timetables found for comparison.')
        return redirect('timetable:department_saved_timetable_list', department_id=department_id)
    
    # Prepare comparison data
    comparison_data = []
    
    for tt in timetables:
        # Extract workload analysis from teacher_year_chart
        workload = tt.teacherYearChart.get('workload_analysis', {})
        thresholds = tt.teacherYearChart.get('thresholds', {})
        teacher_totals = tt.teacherYearChart.get('teacher_totals', {})
        
        # Calculate statistics
        total_lecture_hours = 0
        total_practical_hours = 0
        
        # Count hours from lecture and practical tables
        for year, rows in tt.lectureHoursTable.items():
            for row in rows:
                total_lecture_hours += row.get('assigned', 0)
        
        for year, rows in tt.practicalHoursTable.items():
            for row in rows:
                total_practical_hours += row.get('assigned', 0)
        
        # Teacher distribution
        teacher_count = len(teacher_totals)
        avg_teacher_hours = workload.get('avg_hours', 0)
        overloaded_count = len(workload.get('overloaded', []))
        underutilized_count = len(workload.get('underutilized', []))
        balanced_count = len(workload.get('balanced', []))
        
        # Get all years from display timetable
        years = list(tt.displayTimetable.keys()) if tt.displayTimetable else []
        
        comparison_data.append({
            'timetable': tt,
            'stats': {
                'total_lecture_hours': total_lecture_hours,
                'total_practical_hours': total_practical_hours,
                'total_hours': total_lecture_hours + total_practical_hours,
                'teacher_count': teacher_count,
                'avg_teacher_hours': round(avg_teacher_hours, 1),
                'overloaded_teachers': overloaded_count,
                'underutilized_teachers': underutilized_count,
                'balanced_teachers': balanced_count,
                'years_count': len(years),
                'years': years,
            },
            'workload': workload,
            'thresholds': thresholds,
            'teacher_totals': teacher_totals,
        })
    
    # Find best timetable based on fitness and balance
    best_tt = max(comparison_data, key=lambda x: (
        x['timetable'].fitness or 0,
        x['stats']['balanced_teachers'],
        -x['stats']['overloaded_teachers']
    ))
    
    return render(request, 'timetable/compare_timetables.html', {
        'department': department,
        'comparison_data': comparison_data,
        'best_tt_id': best_tt['timetable'].id,
    })


# ==================== EXPORT VIEWS ====================

def _cell_to_text(cell_data):
    """Convert a display_tt cell dict to a human-readable string for exports."""
    if cell_data is None:
        return "-"
    cell_type = cell_data.get("type")
    if cell_type == "recess":
        return "Recess"
    if cell_type == "lecture":
        details = cell_data.get("details", {})
        parts = []
        if details.get("lecture"):
            parts.append(details["lecture"])
        if details.get("teacher"):
            parts.append(details["teacher"])
        return "\n".join(parts) if parts else "-"
    if cell_type == "practical":
        batches = cell_data.get("batches", [])
        lines = []
        for b in batches:
            lines.append(f"{b.get('batch','')} | {b.get('lab','')} | {b.get('faculty','')}")
        return "\n".join(lines) if lines else "-"
    return "-"


def export_timetable_excel(request, tt_id):
    """Export timetable to Excel format using the display_tt structure."""
    tt = get_object_or_404(DepartmentTimetableModel, pk=tt_id)

    dept = tt.department
    raw_recess_windows = [
        [r.recess_start_time.strftime('%H:%M'), r.recess_end_time.strftime('%H:%M')]
        for r in dept.relatedNameRecesses.order_by('recess_start_time')
    ]
    normalized_recess_windows = normalize_recess_windows(
        raw_recess_windows,
        dept.start_time.strftime('%H:%M'),
        dept.end_time.strftime('%H:%M'),
    )
    department_recess_starts = [start for start, _ in normalized_recess_windows]

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    recess_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    lecture_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    practical_fill = PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=13)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    display_tt = tt.displayTimetable

    for year_name, year_data in display_tt.items():
        ws = wb.create_sheet(title=year_name[:31])

        days = year_data.get("days", [])
        rows = year_data.get("rows", [])

        # Safety for printable output: always include department recess rows.
        # If a recess row exists but contains non-recess cells, force it to recess.
        if days and department_recess_starts:
            rows_by_time = {
                (row.get("time") or ""): {
                    "time": row.get("time"),
                    "cells": list(row.get("cells", [])),
                }
                for row in rows
                if row.get("time")
            }

            for recess_time in department_recess_starts:
                recess_cells = [{"type": "recess"} for _ in days]
                if recess_time in rows_by_time:
                    rows_by_time[recess_time]["cells"] = recess_cells
                else:
                    rows_by_time[recess_time] = {
                        "time": recess_time,
                        "cells": recess_cells,
                    }

            rows = sorted(
                rows_by_time.values(),
                key=lambda r: datetime.strptime(r.get("time", "00:00"), "%H:%M")
            )

        num_cols = len(days) + 2  # Time col + Day cols + 1-based offset

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols - 1, 2))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{tt.department.department_name} — {year_name} — {tt.name}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22

        # Header row (Time | Mon | Tue | ...)
        excel_row = 3
        ws.cell(row=excel_row, column=1, value="Time").font = header_font
        ws.cell(row=excel_row, column=1).fill = header_fill
        ws.cell(row=excel_row, column=1).border = border
        ws.cell(row=excel_row, column=1).alignment = Alignment(horizontal='center')

        for col_idx, day_name in enumerate(days, start=2):
            c = ws.cell(row=excel_row, column=col_idx, value=day_name)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal='center')

        excel_row += 1

        # Data rows
        for timetable_row in rows:
            time_label = timetable_row.get("time", "")
            cells = timetable_row.get("cells", [])

            time_cell = ws.cell(row=excel_row, column=1, value=time_label)
            time_cell.font = Font(bold=True)
            time_cell.border = border
            time_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col_idx, cell_data in enumerate(cells, start=2):
                c = ws.cell(row=excel_row, column=col_idx)
                c.value = _cell_to_text(cell_data)
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if cell_data:
                    ct = cell_data.get("type")
                    if ct == "recess":
                        c.fill = recess_fill
                    elif ct == "lecture":
                        c.fill = lecture_fill
                    elif ct == "practical":
                        c.fill = practical_fill

            ws.row_dimensions[excel_row].height = 40
            excel_row += 1

        # Column widths
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(days) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{tt.name}_timetable.xlsx"'
    wb.save(response)
    return response


def export_timetable_pdf(request, tt_id):
    """Export timetable to PDF format using the display_tt structure."""
    tt = get_object_or_404(DepartmentTimetableModel, pk=tt_id)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{tt.name}_timetable.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A3),
                            leftMargin=0.24*inch, rightMargin=0.24*inch,
                            topMargin=0.3*inch, bottomMargin=0.3*inch)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=20, textColor=colors.HexColor('#1f4788'),
        spaceAfter=6, alignment=1
    )
    heading_style = ParagraphStyle(
        'CustomHeading', parent=styles['Heading2'],
        fontSize=15, textColor=colors.HexColor('#2e5c8a'), spaceAfter=8
    )
    meta_style = ParagraphStyle(
        'PDFMeta', parent=styles['Normal'],
        fontSize=11, textColor=colors.HexColor('#334155'),
        leading=13, alignment=0
    )

    header_cell_style = ParagraphStyle(
        'PDFHeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9.4,
        leading=11,
        alignment=1,
        textColor=colors.whitesmoke,
    )
    time_cell_style = ParagraphStyle(
        'PDFTimeCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=10.6,
        alignment=1,
    )
    body_cell_style = ParagraphStyle(
        'PDFBodyCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.1,
        leading=10.8,
        alignment=0,
        textColor=colors.HexColor('#0f172a'),
    )

    def pdf_paragraph(html_text, style):
        return Paragraph(html_text or '-', style)

    def compact_teacher_name(name):
        raw = str(name or '').replace('.', '. ').strip()
        raw = ' '.join(raw.split())
        if not raw:
            return ''

        parts = raw.split(' ')
        known_prefixes = {'Mr', 'Ms', 'Mrs', 'Dr', 'Prof'}

        prefix = ''
        if parts and parts[0].rstrip('.') in known_prefixes:
            prefix = parts[0].rstrip('.') + '.'
            parts = parts[1:]

        if not parts:
            return prefix
        if len(parts) == 1:
            core = parts[0]
        else:
            core = f"{parts[0][0].upper()}. {parts[-1]}"

        return f"{prefix} {core}".strip()

    def cell_to_pdf_html(cell_data):
        if not cell_data:
            return '-'

        cell_type = cell_data.get('type')

        if cell_type == 'recess':
            return '<b>Recess</b>'

        if cell_type == 'lecture':
            details = cell_data.get('details', {})
            lecture = escape(str(details.get('lecture', '')).strip())
            teacher = escape(compact_teacher_name(details.get('teacher', '')))
            if lecture and teacher:
                return f"<b>{lecture}</b><br/><font color='#475569'>{teacher}</font>"
            if lecture:
                return f"<b>{lecture}</b>"
            if teacher:
                return f"<font color='#475569'>{teacher}</font>"
            return '-'

        if cell_type == 'practical':
            batches = cell_data.get('batches', []) or []
            parts = []
            for b in batches:
                batch = escape(str(b.get('batch', '')).strip())
                lab = escape(str(b.get('lab', '')).strip())
                faculty = escape(compact_teacher_name(b.get('faculty', '')))
                left = ''
                if batch and lab:
                    left = f"<b>{batch}</b> {lab}"
                elif batch:
                    left = f"<b>{batch}</b>"
                elif lab:
                    left = lab

                if left and faculty:
                    parts.append(f"{left}<br/><font color='#475569'>{faculty}</font>")
                elif left:
                    parts.append(left)
                elif faculty:
                    parts.append(f"<font color='#475569'>{faculty}</font>")

            return '<br/><br/>'.join(parts) if parts else '-'

        return '-'

    display_tt = tt.displayTimetable

    elements.append(Paragraph(f"{tt.department.department_name} — {tt.name}", title_style))
    elements.append(Spacer(1, 0.1*inch))
    if tt.fitness:
        elements.append(Paragraph(f"Fitness Score: {tt.fitness:.2f}%", meta_style))
        elements.append(Spacer(1, 0.14*inch))

    dept = tt.department
    recess_windows = [
        [r.recess_start_time.strftime('%H:%M'), r.recess_end_time.strftime('%H:%M')]
        for r in dept.relatedNameRecesses.order_by('recess_start_time')
    ]
    configured_slots = generate_time_range(
        dept.hours_in_lecture,
        dept.start_time.strftime('%H:%M'),
        dept.end_time.strftime('%H:%M'),
        recess_windows,
    )

    year_keys = list(display_tt.keys())

    for idx, (year_name, year_data) in enumerate(display_tt.items()):
        elements.append(Paragraph(f"Year: {year_name}", heading_style))
        elements.append(Spacer(1, 0.08*inch))

        days = year_data.get("days", [])
        rows = year_data.get("rows", [])

        if not days or not rows:
            elements.append(Paragraph("No timetable data available.", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            continue

        # Normalize rows using configured department slots so export always includes
        # every slot defined in setup, even when legacy saved rows are incomplete.
        day_count = len(days)
        slot_row_map = {}
        for row in rows:
            label = str(row.get('time', '')).strip()
            if label:
                slot_row_map[label] = row

        configured_slot_set = set(configured_slots)
        normalized_rows = []
        for slot in configured_slots:
            slot_is_recess = is_recess_time(
                slot,
                recess_windows,
                dept.start_time.strftime('%H:%M'),
                dept.end_time.strftime('%H:%M'),
            )

            if slot_is_recess:
                normalized_rows.append({
                    "time": slot,
                    "cells": [{"type": "recess"} for _ in range(day_count)]
                })
                continue

            source_row = slot_row_map.get(slot)
            if source_row:
                source_cells = list(source_row.get('cells', []) or [])
                if len(source_cells) < day_count:
                    source_cells.extend([None] * (day_count - len(source_cells)))
                elif len(source_cells) > day_count:
                    source_cells = source_cells[:day_count]
                normalized_rows.append({"time": slot, "cells": source_cells})
            else:
                normalized_rows.append({"time": slot, "cells": [None] * day_count})

        # Preserve legacy extra rows that are not part of current configured slots.
        extra_rows = [r for r in rows if str(r.get('time', '')).strip() not in configured_slot_set]
        if extra_rows:
            normalized_rows.extend(extra_rows)

        rows = normalized_rows

        num_cols = len(days) + 1
        time_col_w = 0.62 * inch
        day_col_w = max(1.0 * inch, (doc.width - time_col_w) / max(len(days), 1))
        col_widths = [time_col_w] + [day_col_w] * len(days)

        def build_table_for_rows(rows_subset):
            table_data = [[pdf_paragraph("Time", header_cell_style)] + [pdf_paragraph(escape(day), header_cell_style) for day in days]]
            row_heights = [0.32 * inch]

            recess_rows = []
            lecture_cells = []
            practical_cells = []
            empty_cells = []

            for row_idx, timetable_row in enumerate(rows_subset, start=1):
                time_label = timetable_row.get("time", "")
                cells = timetable_row.get("cells", [])

                time_html = f"<b>{escape(str(time_label or '-'))}</b>"
                pdf_row = [Paragraph(time_html, time_cell_style)]
                max_required_height = 0.48 * inch

                _, time_required_h = pdf_row[0].wrap(col_widths[0] - 8, 1000)
                max_required_height = max(max_required_height, time_required_h + 10)

                is_recess_row = False
                for col_idx, cell_data in enumerate(cells, start=1):
                    cell_html = cell_to_pdf_html(cell_data)
                    cell_paragraph = Paragraph(cell_html, body_cell_style)
                    pdf_row.append(cell_paragraph)

                    _, required_h = cell_paragraph.wrap(col_widths[col_idx] - 8, 1000)
                    max_required_height = max(max_required_height, required_h + 10)

                    if cell_data and cell_data.get("type") == "recess":
                        is_recess_row = True
                    elif cell_data and cell_data.get("type") == "lecture":
                        lecture_cells.append((col_idx, row_idx))
                    elif cell_data and cell_data.get("type") == "practical":
                        practical_cells.append((col_idx, row_idx))
                    elif not cell_data:
                        empty_cells.append((col_idx, row_idx))

                table_data.append(pdf_row)
                row_heights.append(max_required_height)
                if is_recess_row:
                    recess_rows.append(row_idx)

            style_cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#365fa8')),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                ('VALIGN', (1, 1), (-1, -1), 'TOP'),
                ('VALIGN', (0, 1), (0, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#9aa3af')),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.HexColor('#fcfcfd'), colors.HexColor('#f8fafc')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
            ]

            for c, r in lecture_cells:
                style_cmds.append(('BACKGROUND', (c, r), (c, r), colors.HexColor('#eef4ff')))
            for c, r in practical_cells:
                style_cmds.append(('BACKGROUND', (c, r), (c, r), colors.HexColor('#edf9ef')))

            for c, r in empty_cells:
                style_cmds.append(('ALIGN', (c, r), (c, r), 'CENTER'))
                style_cmds.append(('TEXTCOLOR', (c, r), (c, r), colors.HexColor('#64748b')))

            for r in recess_rows:
                style_cmds.append(('BACKGROUND', (1, r), (-1, r), colors.HexColor('#D9D9D9')))
                style_cmds.append(('ALIGN', (1, r), (-1, r), 'CENTER'))
                style_cmds.append(('FONTNAME', (1, r), (-1, r), 'Helvetica-Bold'))
                style_cmds.append(('TEXTCOLOR', (1, r), (-1, r), colors.HexColor('#374151')))

            table = Table(table_data, colWidths=col_widths[:num_cols], rowHeights=row_heights, repeatRows=1)
            table.setStyle(TableStyle(style_cmds))
            return table

        table = build_table_for_rows(rows)
        table.splitByRow = 1
        elements.append(table)
        elements.append(Spacer(1, 0.06*inch))

        if idx < len(year_keys) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    return response


# ============================================================================
# BATCH OPERATIONS VIEWS
# ============================================================================

@require_http_methods(["POST"])
def batch_delete_subjects(request, department_id):
    """Delete multiple subjects at once"""
    try:
        data = json.loads(request.body)
        subject_ids = data.get('ids', [])
        
        if not subject_ids:
            return JsonResponse({
                'success': False,
                'message': 'No subjects selected'
            }, status=400)
        
        # Get department
        department = get_object_or_404(setupModel, pk=department_id)
        
        # Delete subjects
        deleted_count = 0
        for subject_id in subject_ids:
            try:
                subject = subjectModel.objects.get(pk=subject_id, year__department=department)
                subject.delete()
                deleted_count += 1
            except subjectModel.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} subjects'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@require_http_methods(["POST"])
def batch_delete_practicals(request, department_id):
    """Delete multiple practicals at once"""
    try:
        data = json.loads(request.body)
        practical_ids = data.get('ids', [])
        
        if not practical_ids:
            return JsonResponse({
                'success': False,
                'message': 'No practicals selected'
            }, status=400)
        
        # Get department
        department = get_object_or_404(setupModel, pk=department_id)
        
        # Delete practicals
        deleted_count = 0
        for practical_id in practical_ids:
            try:
                practical = practicalModel.objects.get(pk=practical_id, year__department=department)
                practical.delete()
                deleted_count += 1
            except practicalModel.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} practicals'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@require_http_methods(["POST"])
def batch_delete_timetables(request):
    """Delete multiple saved timetables at once"""
    try:
        data = json.loads(request.body)
        timetable_ids = data.get('ids', [])
        
        if not timetable_ids:
            return JsonResponse({
                'success': False,
                'message': 'No timetables selected'
            }, status=400)
        
        # Delete timetables
        deleted_count = 0
        for tt_id in timetable_ids:
            try:
                timetable = DepartmentTimetableModel.objects.get(pk=tt_id)
                timetable.delete()
                deleted_count += 1
            except DepartmentTimetableModel.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} timetables'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@require_http_methods(["POST"])
def batch_delete_years(request, department_id):
    """Delete multiple years at once"""
    try:
        data = json.loads(request.body)
        year_ids = data.get('ids', [])
        
        if not year_ids:
            return JsonResponse({
                'success': False,
                'message': 'No years selected'
            }, status=400)
        
        # Get department
        department = get_object_or_404(setupModel, pk=department_id)
        
        # Delete years
        deleted_count = 0
        for year_id in year_ids:
            try:
                year = yearSetupModel.objects.get(pk=year_id, department=department)
                year.delete()
                deleted_count += 1
            except yearSetupModel.DoesNotExist:
                continue
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} years'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
